"""Export extraction results to CSV / JSON / XLSX.

XLSX adds conditional formatting based on the threshold flags so the buyer
can spot bad readings at a glance.
"""
from __future__ import annotations

import io
import json
from dataclasses import asdict
from datetime import datetime
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.extraction import ExtractionResult, RoomReading
from modules.flags import flag_row, severity_of

CSV_COLUMNS = [
    "room", "day",
    "IN_pH", "IN_EC", "IN_VWC", "IN_solus",
    "S1_pH", "S1_EC", "S1_VWC", "S1_solus",
    "S2_pH", "S2_EC", "S2_VWC", "S2_solus",
    "S3_pH", "S3_EC", "S3_VWC", "S3_solus",
    "S4_pH", "S4_EC", "S4_VWC", "S4_solus",
    "S5_pH", "S5_EC", "S5_VWC", "S5_solus",
    "temp", "humidity", "co2", "notes", "flags",
]

# Excel fill colors per severity
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_ORANGE = PatternFill(start_color="FFD8A8", end_color="FFD8A8", fill_type="solid")
_FONT_BOLD = Font(bold=True)


def _rows_to_dataframe(rows: List[RoomReading]) -> pd.DataFrame:
    """Convert RoomReading list to a DataFrame with stable column order."""
    if not rows:
        return pd.DataFrame(columns=CSV_COLUMNS)
    records = []
    for r in rows:
        d = asdict(r)
        d["flags"] = ";".join(d.get("flags") or [])
        records.append(d)
    df = pd.DataFrame(records)
    for col in CSV_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[CSV_COLUMNS]


def to_csv(result: ExtractionResult) -> bytes:
    """Return CSV bytes (one row per room reading)."""
    df = _rows_to_dataframe(result.rows)
    return df.to_csv(index=False).encode("utf-8")


def to_json(result: ExtractionResult) -> bytes:
    """Return JSON bytes with structured nested metadata."""
    payload: Dict[str, Any] = {
        "extracted_at": datetime.utcnow().isoformat() + "Z",
        "model": result.model,
        "building": result.building,
        "room_override": result.room_override,
        "source_images": result.source_images,
        "rows": [r.to_dict() for r in result.rows],
        "raw_text": result.raw_text,
    }
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def to_xlsx(result: ExtractionResult, thresholds: Dict[str, Any]) -> bytes:
    """Return XLSX bytes with conditional formatting on flagged cells.

    Red fill = high/low/combo. Yellow fill = borderline.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extractions"

    ws.append(CSV_COLUMNS)
    for cell in ws[1]:
        cell.font = _FONT_BOLD

    df = _rows_to_dataframe(result.rows)
    col_index = {name: i + 1 for i, name in enumerate(CSV_COLUMNS)}

    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_name, value in zip(CSV_COLUMNS, row_data):
            ws.cell(row=row_idx, column=col_index[col_name], value=value)

        # Apply flags after the row is written.
        original = result.rows[row_idx - 2]
        cell_flags = flag_row(original.to_dict(), thresholds)
        for field_name, flags in cell_flags.items():
            if not flags:
                continue
            sev = severity_of(flags)
            col_letter = get_column_letter(col_index[field_name])
            cell = ws[f"{col_letter}{row_idx}"]
            if sev == "combo":
                cell.fill = _FILL_ORANGE
                cell.font = _FONT_BOLD
            elif sev in ("high", "low"):
                cell.fill = _FILL_RED
            elif sev == "borderline":
                cell.fill = _FILL_YELLOW

    # Best-fit column widths.
    for col_idx, name in enumerate(CSV_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, len(name) + 2)

    # Second sheet: extraction metadata
    meta = wb.create_sheet("Metadata")
    meta.append(["Field", "Value"])
    meta["A1"].font = _FONT_BOLD
    meta["B1"].font = _FONT_BOLD
    meta.append(["Extracted at", datetime.utcnow().isoformat() + "Z"])
    meta.append(["Model", result.model])
    meta.append(["Building", result.building or ""])
    meta.append(["Room override", result.room_override or ""])
    meta.append(["Source images", ", ".join(result.source_images)])
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
